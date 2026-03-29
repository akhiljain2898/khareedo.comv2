"""
worker/xlsx_builder.py
openpyxl workbook builder.
Single tab, 6 columns, in-memory — never written to disk.
Uploads two copies to R2:
  - download/{order_id}.xlsx  (customer-facing, 30 min TTL via R2 lifecycle)
  - archive/{order_id}.xlsx   (ops-only, 7 day TTL via R2 lifecycle)
"""

import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from common.r2_client import upload_bytes
from common.config import TARGET_RESULT_COUNT

logger = logging.getLogger(__name__)

# Column definitions: (header label, contact dict key, column width)
COLUMNS = [
    ("Company Name",        "name",                36),
    ("Phone",               "phone",               20),
    ("Email",               "email",               30),
    ("Address",             "address",             36),
    ("Product Description", "product_description", 50),
    ("Website",             "website",             40),
]

# Header styling — navy background matching khareedo brand
HEADER_FILL  = PatternFill(start_color="1A2E4A", end_color="1A2E4A", fill_type="solid")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")


def _build_workbook(results: list[dict], product_name: str) -> bytes:
    """
    Build the Excel workbook in memory and return raw bytes.
    Never touches disk.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Worksheet metadata
    ws.sheet_properties.tabColor = "1A2E4A"

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"  # Freeze header row

    # Data rows
    for row_idx, contact in enumerate(results, start=2):
        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = contact.get(key) or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row_idx].height = 30

    # If partial results, add a note at the bottom
    if len(results) < TARGET_RESULT_COUNT:
        note_row = len(results) + 3
        note_cell = ws.cell(
            row=note_row,
            column=1,
            value=(
                f"We found {len(results)} verified suppliers for '{product_name}'. "
                f"If you need more results, email admin@khareedo.com with your Order ID."
            )
        )
        note_cell.font = Font(name="Calibri", size=10, italic=True, color="8A9AB0")
        ws.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row, end_column=len(COLUMNS)
        )

    # Serialize to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def build_and_upload(order_id: str, product_name: str, results: list[dict]) -> str:
    """
    Build Excel workbook and upload both copies to R2.
    
    Returns:
        archive_key: the R2 key for the archive copy (stored in Postgres)
    
    R2 lifecycle rules (set in Cloudflare dashboard):
        - download/* → delete after 1800 seconds (30 min)
        - archive/*  → delete after 7 days
    """
    logger.info(f"Building Excel for order {order_id} — {len(results)} rows")

    xlsx_bytes = _build_workbook(results, product_name)

    download_key = f"download/{order_id}.xlsx"
    archive_key  = f"archive/{order_id}.xlsx"

    # Upload customer-facing copy
    upload_bytes(download_key, xlsx_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    logger.info(f"Uploaded to R2: {download_key}")

    # Upload ops archive copy (same bytes)
    upload_bytes(archive_key, xlsx_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    logger.info(f"Uploaded to R2: {archive_key}")

    return archive_key
